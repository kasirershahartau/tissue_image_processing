import numpy as np
import pandas as pd
import scipy as sp
import os
import matplotlib
from matplotlib import pyplot as plt
import statsmodels.api as sm
import statsmodels.formula.api as smf
import scipy.stats as stats
from scipy.stats import norm, chi2, shapiro, skew, mannwhitneyu
from scipy.interpolate import interp1d
from statsmodels.formula.api import ols
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import scikit_posthocs as sp
from statsmodels.stats.multitest import multipletests
from openpyxl import load_workbook

class DataCollector:
    def __init__(self, name, folders=[], file_names=[], data_labels=[], normalization=1, sample=None, omit_zeros=False):
        self.name = name
        self.normalization = normalization
        self.labels = data_labels
        self.experiment_idx = None
        self.files = None
        self.initial_batch_indices = None
        if sample is not None:
            self.sample = sample/normalization
        else:
            self.experiment_idx = [folders.index(folder) for folder in folders]
            self.files = [os.path.join(folder, file_name) for folder, file_name in zip(folders, file_names)]
            self.sample = self.collect(omit_zeros=omit_zeros)

    def collect(self, omit_zeros=False):
        s = np.empty(0)
        self.initial_batch_indices = []
        group_id = 0
        for f, l in zip(self.files, self.labels):
            all_data = pd.read_pickle(f)
            relevant_data = all_data[l].to_numpy()
            if omit_zeros:
                relevant_data = relevant_data[relevant_data != 0]
            self.initial_batch_indices.append(s.size)
            if hasattr(self.normalization, "__len__"):
                normalization = self.normalization[group_id]
            else:
                normalization = self.normalization
            s = np.hstack((s, relevant_data/ normalization))
            group_id += 1
        self.initial_batch_indices = np.array(self.initial_batch_indices)
        return s[~np.isnan(s)]

    def get_name(self):
        return self.name

    def get_sample(self):
        return self.sample

    def get_sample_size(self):
        return self.sample.size

    def get_partial_sample_size(self, file_index):
        start = self.initial_batch_indices[file_index]
        end = self.sample.size if file_index + 1 >= self.initial_batch_indices.size else self.initial_batch_indices[
            file_index + 1]
        return end - start

    def get_partial_sample(self, file_index):
        if self.initial_batch_indices is None:
            return self.sample
        start = self.initial_batch_indices[file_index]
        end = self.sample.size if file_index + 1 >= self.initial_batch_indices.size  else self.initial_batch_indices[file_index + 1]
        return self.sample[start:end]

    def get_biological_repeat(self, file_index):
        if self.experiment_idx is None:
            return 0
        return self.experiment_idx[file_index]

    def get_number_of_data_points(self):
        return self.sample.size

    def get_average(self):
        return np.average(self.sample)

    def get_group_avg(self, group_id=-1):
        if group_id >= 0:
            return np.average(self.get_partial_sample(group_id))
        else:
            return [np.average(self.get_partial_sample(i)) for i in range(self.get_number_of_groups())]

    def get_average_of_groups(self):
        return np.average(self.get_group_avg())

    def get_std_of_groups(self):
        return np.std(self.get_group_avg())

    def get_se_of_groups(self):
        if self.get_std_of_groups() > 1:
            return self.get_std_of_groups()/np.sqrt(self.get_number_of_groups())
        else:
            return self.get_se()

    def get_number_of_groups(self):
        if self.files is None:
            return 1
        return len(self.files)

    def get_std(self):
        return np.std(self.sample)

    def get_se(self):
        return self.get_std() / np.sqrt(self.get_number_of_data_points())

    def get_group_std(self):
        avg = np.array([self.get_group_avg(i) for i in range(self.get_number_of_groups())])
        return np.std(avg)

    def get_group_se(self):
        return self.get_group_std()/np.sqrt(self.get_number_of_groups())

    def get_max(self):
        return np.max(self.sample)

    def get_min(self):
        return np.min(self.sample)

    def save_sample(self, out_path, by_groups=False):
        if by_groups:
            for group in range(self.get_number_of_groups()):
                group_sample = self.get_partial_sample(group)
                np.save(os.path.join(out_path, "%s_experiment%d.npy" % (self.name, group)), group_sample)
        else:
            np.save(os.path.join(out_path, "%s.npy" % self.name), self.sample)

    def save_to_excel(self, out_path, data_label, change_to_int=False):
        df = pd.DataFrame()
        for group in range(self.get_number_of_groups()):
            group_sample = self.get_partial_sample(group)
            if change_to_int:
                group_sample = group_sample.astype(int)
            current_df = pd.DataFrame({"Experiment #": [group]*group_sample.size, "Cell #":np.arange(group_sample.size),
                                       data_label: group_sample})
            df = pd.concat([df, current_df], ignore_index=True)
        if os.path.isfile(out_path):
            mode = "a"
        else:
            mode = "w"
        with pd.ExcelWriter(out_path, mode=mode) as writer:
            df.to_excel(writer, sheet_name=self.name[:30])
        return df

    def divide(self, other):
        sample1 = self.sample
        sample2 = other.sample
        new_initial_batch_indices = []
        zeros = (sample2 == 0)
        cum_zeros = np.cumsum(zeros.astype(int))
        for i in self.initial_batch_indices:
             if i > 0:
                new_initial_batch_indices.append(i - cum_zeros[i-1])
             else:
                 new_initial_batch_indices.append(i)
        self.sample = sample1[sample2 != 0] / sample2[sample2 != 0]
        self.name = self.name + " per " + other.name


def _append_row_to_excel(filename, sheet_name, row_dict):
    """
    Appends a single row (dict) to an Excel sheet.
    Creates the file or sheet if needed.
    """

    # -----------------------------
    # CASE 1 — File does NOT exist
    # -----------------------------
    if not os.path.exists(filename):
        df = pd.DataFrame([row_dict])
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # -----------------------------
    # CASE 2 — File exists
    # -----------------------------
    try:
        book = load_workbook(filename)
    except Exception:
        # File exists but is corrupted or not a real Excel file
        print(f"Warning: '{filename}' is not a valid Excel file. Recreating it.")
        df = pd.DataFrame([row_dict])
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # -----------------------------
    # CASE 2A — Sheet exists → append
    # -----------------------------
    if sheet_name in book.sheetnames:
        existing = pd.read_excel(filename, sheet_name=sheet_name)
        new_df = pd.concat([existing, pd.DataFrame([row_dict])], ignore_index=True)

        with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            new_df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # -----------------------------
    # CASE 2B — Sheet does NOT exist → add new sheet
    # -----------------------------
    # Load all existing sheets
    all_sheets = {name: pd.read_excel(filename, sheet_name=name) for name in book.sheetnames}

    # Add the new sheet
    all_sheets[sheet_name] = pd.DataFrame([row_dict])

    # Rewrite the entire file with all sheets
    with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
        for name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

class TwoSampleCompare:
    def __init__(self, sample1, sample2, sample1_label='sample1', sample2_label='sample2', continues=True):
        self.sample1 = np.array(sample1)
        self.sample2 = np.array(sample2)
        self.sample1_label = sample1_label
        self.sample2_label = sample2_label
        self.continues = continues

    # ---------------------------------------------------------
    # NORMALITY CHECK
    # ---------------------------------------------------------

    @staticmethod
    def is_normal(sample):
        sample = np.array(sample)

        # Too small for reliable normality tests
        if sample.size < 5:
            return False

        # Shapiro–Wilk (best for small n)
        if sample.size < 20:
            p = stats.shapiro(sample).pvalue
            return p > 0.05

        # D’Agostino–Pearson (better for n >= 20)
        p = stats.normaltest(sample).pvalue
        return p > 0.05

    # ---------------------------------------------------------
    # VARIANCE EQUALITY CHECK
    # ---------------------------------------------------------

    @staticmethod
    def equal_variances(x, y):
        # Levene’s test (robust to non-normality)
        p = stats.levene(x, y).pvalue
        return p > 0.05

    # ---------------------------------------------------------
    # MAIN COMPARISON
    # ---------------------------------------------------------

    def compare_samples(self, verbose=False, save_to_excel=None, sheet="", label=""):

        # If data is not continues → always use Mann–Whitney
        if not self.continues:
            stat, p = stats.mannwhitneyu(self.sample1, self.sample2, alternative='two-sided')
            if verbose:
                print(f"Using Mann–Whitney U test: U={stat:.3f}, p={p:.3f}")

            results = {
                "test": "Mann–Whitney U",
                "stat": stat,
                "p": p,
                "normal1": False,
                "normal2": False,
                "normal": False,
                "equal_var": None,
            }

        # Check normality
        normal1 = self.is_normal(self.sample1)
        normal2 = self.is_normal(self.sample2)
        normal = normal1 and normal2

        if not normal:
            stat, p = stats.mannwhitneyu(self.sample1, self.sample2, alternative='two-sided')
            if verbose:
                print(f"Using Mann–Whitney U test (non-normal data): U={stat:.3f}, p={p:.3f}")

            results = {
                "test": "Mann–Whitney U",
                "stat": stat,
                "p": p,
                "normal1": normal1,
                "normal2": normal2,
                "normal": normal,
                "equal_var": None,
            }
        else:
            # Normal → use t-test
            equal_var = self.equal_variances(self.sample1, self.sample2)
            test_name = "Student's t-test" if equal_var else "Welch's t-test"

            stat, p = stats.ttest_ind(self.sample1, self.sample2, equal_var=equal_var)
            if verbose:
                print(f"Using {test_name}: t={stat:.3f}, p={p:.3f}")

            results = {
                "test": test_name,
                "stat": stat,
                "p": p,
                "normal1": normal1,
                "normal2": normal2,
                "normal": normal,
                "equal_var": equal_var,
            }

        if save_to_excel:
            if save_to_excel:
                row = {
                    "label": label,
                    "sample1": self.sample1_label,
                    "sample2": self.sample2_label,
                    "test_used": results["test"],
                    "statistic": results["stat"],
                    "p_value": results["p"],
                    "normality_sample1": results["normal1"],
                    "normality_sample2": results["normal2"],
                    "both_normal": results["normal"],
                    "equal_variances": results["equal_var"],
                }
                _append_row_to_excel(save_to_excel, sheet, row)

        return p

class TwoByTwoCompare:
    """
    Compare four samples in a 2×2 factorial design:
    Factor A: two levels (A1, A2)
    Factor B: two levels (B1, B2)

    Samples must be provided as:
        A1B1, A1B2, A2B1, A2B2
    """

    def __init__(self,
                 A1B1, A1B2, A2B1, A2B2,
                 factorA_name="FactorA",
                 factorB_name="FactorB",
                 A_levels=("A1", "A2"),
                 B_levels=("B1", "B2"),
                 continues=True):

        self.samples = {
            (A_levels[0], B_levels[0]): np.array(A1B1),
            (A_levels[0], B_levels[1]): np.array(A1B2),
            (A_levels[1], B_levels[0]): np.array(A2B1),
            (A_levels[1], B_levels[1]): np.array(A2B2),
        }

        self.factorA_name = factorA_name
        self.factorB_name = factorB_name
        self.A_levels = A_levels
        self.B_levels = B_levels
        self.continues = continues

        # Build long-format dataframe for ANOVA
        self.df = self._build_dataframe()

    # ---------------------------------------------------------
    # DATAFRAME BUILDER
    # ---------------------------------------------------------

    def _build_dataframe(self):
        rows = []
        for (A, B), values in self.samples.items():
            for v in values:
                rows.append({self.factorA_name: A,
                             self.factorB_name: B,
                             "value": v})
        return pd.DataFrame(rows)

    # ---------------------------------------------------------
    # NORMALITY CHECK
    # ---------------------------------------------------------

    @staticmethod
    def is_normal(sample):
        sample = np.array(sample)
        if sample.size < 5:
            return False
        if sample.size < 20:
            return stats.shapiro(sample).pvalue > 0.05
        return stats.normaltest(sample).pvalue > 0.05

    # ---------------------------------------------------------
    # VARIANCE HOMOGENEITY CHECK
    # ---------------------------------------------------------

    def equal_variances(self):
        groups = [
            group["value"].to_numpy()
            for _, group in self.df.groupby([self.factorA_name, self.factorB_name])
        ]
        stat, p = stats.levene(*groups)
        return p > 0.05

    def _flatten_posthoc(self, posthoc):
        """
        Convert a posthoc DataFrame into a flat dict of columns.
        Example:
            comparison | p_raw | p_corrected
            A vs B     | 0.01  | 0.02
        becomes:
            posthoc_A_vs_B_p_raw = 0.01
            posthoc_A_vs_B_p_corrected = 0.02
        """

        flat = {}

        if posthoc is None:
            return flat

        df = posthoc.copy()

        # Ensure comparison column exists
        if "comparison" in df.columns:
            comp_col = "comparison"
        else:
            # Tukey summary table → convert to DataFrame
            df = df.reset_index()
            comp_col = df.columns[0]

        for _, row in df.iterrows():
            comp = str(row[comp_col]).replace(" ", "").replace("-", "_").replace(":", "_")
            for col in df.columns:
                if col == comp_col:
                    continue
                key = f"posthoc_{comp}_{col}"
                flat[key] = row[col]

        return flat

    # ---------------------------------------------------------
    # MAIN COMPARISON
    # ---------------------------------------------------------

    def compare(self, verbose=False, save_to_excel=None, sheet="", label=""):
        """
        Performs:
        - Two-way ANOVA if assumptions met
        - Otherwise Scheirer–Ray–Hare nonparametric test
        - Tukey or nonparametric posthoc tests
        - Optionally saves results to Excel (single sheet)
        """

        # Non-continues → Kruskal-based factorial test
        if not self.continues:
            if verbose:
                print("Using Scheirer–Ray–Hare test (non-parametric 2×2 ANOVA).")
            results = self._scheirer_ray_hare(verbose)
        else:
            # Check assumptions
            normal = all(self.is_normal(v) for v in self.samples.values())
            equal_var = self.equal_variances()

            self.normal = normal
            self.equal_var = equal_var

            if verbose:
                print(f"Normality: {normal}, Equal variances: {equal_var}")

            if normal and equal_var:
                results = self._two_way_anova(verbose)
            else:
                if verbose:
                    print("Assumptions violated → using Scheirer–Ray–Hare test.")
                results = self._scheirer_ray_hare(verbose)

        # -----------------------------
        # SAVE TO EXCEL (single sheet)
        # -----------------------------
        if save_to_excel:

            # Determine test type and p-values
            if "anova" in results:
                test_type = "ANOVA"
                pA = results["anova"].loc[self.factorA_name, "PR(>F)"]
                pB = results["anova"].loc[self.factorB_name, "PR(>F)"]
                pAB = results["anova"].loc[f"C({self.factorA_name}):C({self.factorB_name})", "PR(>F)"]

                # Tukey posthoc
                tukey = results["posthoc"]
                # Convert Tukey summary to DataFrame
                tukey_df = pd.DataFrame(tukey._results_table.data[1:], columns=tukey._results_table.data[0])
                posthoc_flat = self._flatten_posthoc(tukey_df)

            else:
                test_type = "SRH"
                pA = results["SRH"].loc[self.factorA_name, "p"]
                pB = results["SRH"].loc[self.factorB_name, "p"]
                pAB = results["SRH"].loc["Interaction", "p"]

                posthoc = results["posthoc"]["posthoc"]
                posthoc_flat = self._flatten_posthoc(posthoc)

            # Base row
            row = {
                "label": label,
                "test_type": test_type,
                "factorA": self.factorA_name,
                "factorB": self.factorB_name,
                "p_factorA": pA,
                "p_factorB": pB,
                "p_interaction": pAB,
            }

            # Merge flattened posthoc columns
            row.update(posthoc_flat)

            _append_row_to_excel(save_to_excel, sheet, row)

        return results

    # ---------------------------------------------------------
    # TWO-WAY ANOVA
    # ---------------------------------------------------------

    def _two_way_anova(self, verbose):
        model = ols(f"value ~ C({self.factorA_name}) * C({self.factorB_name})",
                    data=self.df).fit()
        anova_table = sm.stats.anova_lm(model, typ=2)

        if verbose:
            print("\nTwo-way ANOVA:")
            print(anova_table)

        # Posthoc Tukey
        tukey = pairwise_tukeyhsd(self.df["value"],
                                  self.df[self.factorA_name] + "_" +
                                  self.df[self.factorB_name])

        if verbose:
            print("\nPosthoc Tukey HSD:")
            print(tukey)

        return {"anova": anova_table, "posthoc": tukey}

    # ---------------------------------------------------------
    # SCHEIRER–RAY–HARE (nonparametric 2×2 ANOVA)
    # ---------------------------------------------------------

    def _scheirer_ray_hare(self, verbose):
        df = self.df.copy()
        df["rank"] = stats.rankdata(df["value"])

        A = df.groupby(self.factorA_name)["rank"].sum()
        B = df.groupby(self.factorB_name)["rank"].sum()
        AB = df.groupby([self.factorA_name, self.factorB_name])["rank"].sum()

        N = len(df)
        a = len(self.A_levels)
        b = len(self.B_levels)

        H_A = 12 / (N * (N + 1)) * sum(A ** 2 / df[self.factorA_name].value_counts()) - 3 * (N + 1)
        H_B = 12 / (N * (N + 1)) * sum(B ** 2 / df[self.factorB_name].value_counts()) - 3 * (N + 1)
        H_AB = 12 / (N * (N + 1)) * sum(AB ** 2 / df.groupby([self.factorA_name,
                                                              self.factorB_name]).size()) - H_A - H_B - 3 * (N + 1)

        p_A = 1 - stats.chi2.cdf(H_A, a - 1)
        p_B = 1 - stats.chi2.cdf(H_B, b - 1)
        p_AB = 1 - stats.chi2.cdf(H_AB, (a - 1) * (b - 1))

        results = pd.DataFrame({
            "H": [H_A, H_B, H_AB],
            "p": [p_A, p_B, p_AB]
        }, index=[self.factorA_name, self.factorB_name, "Interaction"])

        # run post-hoc logic
        posthoc = self._posthoc_srh(results)

        if verbose:
            print("\nScheirer–Ray–Hare test:")
            print(results)
            print("\nPosthoc:")
            print(posthoc["type"])
            print(posthoc['posthoc'].to_string())

        return {"SRH": results, "posthoc": posthoc}

    def _posthoc_srh(self, srh_results):
        """
        Post-hoc logic for Scheirer–Ray–Hare test.
        - If interaction significant → simple effects (Mann–Whitney)
        - Else if main effects significant → Dunn test
        """

        pA = srh_results.loc[self.factorA_name, "p"]
        pB = srh_results.loc[self.factorB_name, "p"]
        pAB = srh_results.loc["Interaction", "p"]

        # Build group labels
        df = self.df.copy()
        df["group"] = df[self.factorA_name] + "_" + df[self.factorB_name]

        # Extract the four groups
        g = {}
        for (A, B), vals in self.samples.items():
            g[f"{A}_{B}"] = np.array(vals)

        # ------------------------------------------
        # Interaction significant → simple effects
        # ------------------------------------------
        if pAB < 0.05:
            comparisons = [
                ("A1B1 vs A1B2", g[f"{self.A_levels[0]}_{self.B_levels[0]}"],
                 g[f"{self.A_levels[0]}_{self.B_levels[1]}"]),
                ("A2B1 vs A2B2", g[f"{self.A_levels[1]}_{self.B_levels[0]}"],
                 g[f"{self.A_levels[1]}_{self.B_levels[1]}"]),
                ("A1B1 vs A2B1", g[f"{self.A_levels[0]}_{self.B_levels[0]}"],
                 g[f"{self.A_levels[1]}_{self.B_levels[0]}"]),
                ("A1B2 vs A2B2", g[f"{self.A_levels[0]}_{self.B_levels[1]}"],
                 g[f"{self.A_levels[1]}_{self.B_levels[1]}"]),
            ]

            names = []
            pvals = []

            for name, x, y in comparisons:
                stat, p = mannwhitneyu(x, y, alternative="two-sided")
                names.append(name)
                pvals.append(p)

            # Holm correction
            _, pvals_corr, _, _ = multipletests(pvals, method="holm")

            posthoc = pd.DataFrame({
                "comparison": names,
                "p_raw": pvals,
                "p_corrected": pvals_corr
            })
            return {"type": "simple_effects", "posthoc": posthoc}

        # ------------------------------------------
        # No interaction, but main effect significant → Dunn test
        # ------------------------------------------
        else:
            dunn = sp.posthoc_dunn(df, val_col="value", group_col="group", p_adjust="holm")
            return {"type": "dunn", "posthoc": dunn}


class HierarchicalTwoSamplesCompare:

    def __init__(self, data1, data2, continues=True):
        """
        continues = False  → use Poisson/NB/ZIP/ZINB pipeline
        continues = True → use continues-data pipeline
        """
        self.data = self.rearrange_data_into_table(data1, data2)
        self.continues = continues

    @staticmethod
    def rearrange_data_into_table(data1, data2):
        df = []
        for data_index, data in enumerate([data1, data2]):
            if isinstance(data, DataCollector):
                for group in range(data.get_number_of_groups()):
                    for measurement in data.get_partial_sample(group):
                        df.append({
                            'measurement': measurement,
                            'stage': data_index,
                            'replicate': f"R{data.get_biological_repeat(group)}"
                        })
            elif isinstance(data, list):
                for group in range(len(data)):
                    for measurement in data[group]:
                        df.append({
                            'measurement': measurement,
                            'stage': data_index,
                            'replicate': f"R{group}"
                        })
        df = pd.DataFrame(df)
        df['stage'] = df['stage'].astype('category')
        df['replicate'] = df['replicate'].astype('category')
        return df

    # ---------------------------------------------------------
    # COUNT-DATA DIAGNOSTICS
    # ---------------------------------------------------------

    def check_overdispersion(self, model):
        pearson = sum(model.resid_pearson**2)
        df = model.df_resid
        return pearson / df

    def check_zero_inflation(self):
        y = self.data['measurement']
        lam = y.mean()
        expected_zeros = np.exp(-lam)
        observed_zeros = (y == 0).mean()
        return observed_zeros > expected_zeros * 1.5

    # ---------------------------------------------------------
    # COUNT-DATA MODELS
    # ---------------------------------------------------------

    def fit_poisson(self):
        return smf.glm("measurement ~ stage", data=self.data,
                       family=sm.families.Poisson()).fit()

    def fit_nb(self):
        return smf.glm("measurement ~ stage", data=self.data,
                       family=sm.families.NegativeBinomial()).fit()

    def fit_zip(self):
        import statsmodels.discrete.count_model as cm
        return cm.ZeroInflatedPoisson.from_formula(
            "measurement ~ stage", self.data, exog_infl="stage"
        ).fit()

    def fit_zinb(self):
        import statsmodels.discrete.count_model as cm
        return cm.ZeroInflatedNegativeBinomialP.from_formula(
            "measurement ~ stage", self.data, exog_infl="stage"
        ).fit()

    # ---------------------------------------------------------
    # CONTINUES-DATA MODELS
    # ---------------------------------------------------------

    def fit_lmm(self):
        return smf.mixedlm("measurement ~ stage", self.data,
                           groups=self.data["replicate"]).fit()

    def fit_log_lmm(self):
        self.data["log_measurement"] = np.log(self.data["measurement"])
        return smf.mixedlm("log_measurement ~ stage", self.data,
                           groups=self.data["replicate"]).fit()

    def fit_gamma_glmm(self):
        return smf.glm("measurement ~ stage", data=self.data,
                       family=sm.families.Gamma()).fit()

    def fit_invgauss_glmm(self):
        return smf.glm("measurement ~ stage", data=self.data,
                       family=sm.families.InverseGaussian()).fit()

    # ---------------------------------------------------------
    # MAIN LOGIC
    # ---------------------------------------------------------

    def compare_samples(self, verbose=False, save_to_excel=None, sheet="", label=""):

        # ---------------- COUNT DATA PIPELINE ----------------
        if not self.continues:

            poisson = self.fit_poisson()
            overdisp = self.check_overdispersion(poisson)
            zero_inf = self.check_zero_inflation()

            if zero_inf:
                if overdisp > 1.5:
                    model = self.fit_zinb()
                    model_type = "Zero-Inflated Negative Binomial"
                else:
                    model = self.fit_zip()
                    model_type = "Zero-Inflated Poisson"
            else:
                if overdisp > 1.5:
                    model = self.fit_nb()
                    model_type = "Negative Binomial"
                else:
                    model = poisson
                    model_type = "Poisson"

            p_val = model.pvalues["stage[T.1]"]

            if verbose:
                print(f"\nSelected model: {model_type}\n")
                print(model.summary())
                summary_text = model.summary().as_text()
            else:
                print(p_val)
                summary_text = None

            results = {
                "model_type": model_type,
                "p_value": p_val,
                "overdispersion": overdisp,
                "zero_inflation": zero_inf,
                "normality_p": None,
                "skewness": None,
                "summary": summary_text,
            }

        # ---------------- CONTINUES DATA PIPELINE ----------------
        else:
            y = self.data["measurement"]

            stat, p_norm = shapiro(y)
            sk = skew(y)

            if p_norm > 0.05 and abs(sk) < 1:
                model = self.fit_lmm()
                model_type = "Linear Mixed Model"
                p_val = model.pvalues["stage[T.1]"]

            elif (y > 0).all():
                model = self.fit_log_lmm()
                model_type = "Log-Transformed LMM"
                p_val = model.pvalues["stage[T.1]"]

            else:
                if sk > 2:
                    model = self.fit_invgauss_glmm()
                    model_type = "Inverse Gaussian GLMM"
                else:
                    model = self.fit_gamma_glmm()
                    model_type = "Gamma GLMM"

                p_val = model.pvalues["stage[T.1]"]

            if verbose:
                print(f"\nSelected model: {model_type}\n")
                print(model.summary())
                summary_text = model.summary().as_text()
            else:
                print(p_val)
                summary_text = None

            results = {
                "model_type": model_type,
                "p_value": p_val,
                "overdispersion": None,
                "zero_inflation": None,
                "normality_p": p_norm,
                "skewness": sk,
                "summary": summary_text,
            }
        if save_to_excel:
            row = {
                "label": label,
                "model_type": results["model_type"],
                "p_value": results["p_value"],
                "overdispersion": results.get("overdispersion"),
                "zero_inflation": results.get("zero_inflation"),
                "normality_p": results.get("normality_p"),
                "skewness": results.get("skewness"),
            }
            _append_row_to_excel(save_to_excel, sheet, row)

        return p_val

    def _save_results_to_excel(self, results, filename, label=""):
        """
        Save hierarchical two-sample comparison results to an Excel file.
        Appends if file exists.
        """

        mode = "a" if os.path.exists(filename) else "w"

        with pd.ExcelWriter(filename, engine="openpyxl", mode=mode) as writer:
            # Diagnostics sheet
            diag_df = pd.DataFrame({
                "overdispersion": [results.get("overdispersion")],
                "zero_inflation": [results.get("zero_inflation")],
                "normality_p": [results.get("normality_p")],
                "skewness": [results.get("skewness")],
            })
            diag_df.to_excel(writer, sheet_name=label + " diagnostics", index=False)

            # Model info sheet
            model_df = pd.DataFrame({
                "model_type": [results["model_type"]],
                "p_value": [results["p_value"]],
            })
            model_df.to_excel(writer, sheet_name=label + " model_info", index=False)

            # Model summary (if available)
            if "summary" in results and results["summary"] is not None:
                summary_text = results["summary"]
                summary_df = pd.DataFrame({"summary": summary_text.split("\n")})
                summary_df.to_excel(writer, sheet_name=label + " model_summary", index=False)


def barplot_annotate_brackets(num1, num2, data, center, height, yerr=None, dh=.05, barh=.05, fs=None, maxasterix=None, ax=None):
    """
    Annotate barplot with p-values.

    :param num1: number of left bar to put bracket over
    :param num2: number of right bar to put bracket over
    :param data: string to write or number for generating asterixes
    :param center: centers of all bars (like plt.bar() input)
    :param height: heights of all bars (like plt.bar() input)
    :param yerr: yerrs of all bars (like plt.bar() input)
    :param dh: height offset over bar / bar + yerr in axes coordinates (0 to 1)
    :param barh: bar height in axes coordinates (0 to 1)
    :param fs: font size
    :param maxasterix: maximum number of asterixes to write (for very small p-values)
    """

    if type(data) is str:
        text = data
    else:
        text = str(data)
        # * is p < 0.05
        # ** is p < 0.005
        # *** is p < 0.0005
        # etc.
        # text = ''
        # p = .05
        #
        # while data < p:
        #     text += '*'
        #     p /= 10.
        #
        #     if maxasterix and len(text) == maxasterix:
        #         break
        #
        # if len(text) == 0:
        #     text = 'n. s.'

    lx, ly = center[num1], height[num1]
    rx, ry = center[num2], height[num2]

    if yerr.size:
        ly += yerr[num1]
        ry += yerr[num2]

    ax_y0, ax_y1 = plt.gca().get_ylim()
    dh *= (ax_y1 - ax_y0)
    barh *= (ax_y1 - ax_y0)

    y = max(ly, ry) + dh

    barx = [lx, lx, rx, rx]
    bary = [y, y+barh, y+barh, y]
    mid = ((lx+rx)/2, y+barh)

    ax.plot(barx, bary, c='black')

    kwargs = dict(ha='center', va='bottom')
    if fs is not None:
        kwargs['fontsize'] = fs

    ax.text(*mid, text, **kwargs)
    return (y+barh)/(ax_y1 - ax_y0)

def compare_and_plot_samples(samples_list, pairs_to_compare, continues=True, plot_style="violin", color='white',
                             edge_color='grey', fig=None, ax=None, show_statistics=False, show_N=False,
                             hirarchical=False, scatter=False, hatch=None, save_to_excel=None, excel_sheet=""):

    # Statistical analysis for each pair
    pvalues = np.zeros((len(pairs_to_compare),))
    for index, pair in enumerate(pairs_to_compare):
        sample1_index, sample2_index = pair
        if hasattr(continues, "__len__"):
            continues_sample = continues[sample1_index] and continues[sample2_index]
        else:
            continues_sample = continues
        if hirarchical:
            analyzer = HierarchicalTwoSamplesCompare(samples_list[sample1_index], samples_list[sample2_index],
                                                     continues=continues_sample)
        else:
            analyzer = TwoSampleCompare(samples_list[sample1_index].get_sample(), samples_list[sample2_index].get_sample(),
                                        samples_list[sample1_index].name, samples_list[sample2_index].name,
                                        continues=continues_sample)
        excel_label = samples_list[sample1_index].name + " vs " + samples_list[sample2_index].name
        pvalues[index] = analyzer.compare_samples(save_to_excel=save_to_excel, sheet=excel_sheet, label=excel_label)

    averages = np.array([sample.get_average_of_groups() for sample in samples_list])
    standard_errors = np.array([sample.get_se_of_groups() for sample in samples_list])
    sample_sizes = np.array([sample.get_number_of_data_points() for sample in samples_list])
    labels = [sample.name for sample in samples_list]

    # Creating box plot
    w = 0.6  # bar width
    x = np.arange(len(samples_list))  # x-coordinates of your bars
    if fig is None:
        fig, ax = plt.subplots()
    errors = False
    if not isinstance(color, list):
        color = [color for i in range(len(samples_list))]
    if not isinstance(edge_color, list):
        edge_color = [edge_color for i in range(len(samples_list))]
    if plot_style == "bar":
        ax.bar(x,
               height=averages,
               capsize=12,  # error bar cap width in points
               width=w,  # bar width
               tick_label=labels,
               color=color,  # face color transparent
               edgecolor=edge_color,
               )
        for pos, y, err, color in zip(x,averages, standard_errors, edge_color):
            ax.errorbar(pos, y, err, lw=2, capsize=15, capthick=2, color=color)
    elif plot_style == "box":
        parts = ax.boxplot([s.get_sample() for s in samples_list],
                   vert=True,
                   showmeans=True,
                   labels=labels,
                   positions=x,
                   showcaps=False,
                   patch_artist=True,
                   showfliers=True,
                   )
        for pc, c, ec in zip(parts['boxes'], color, edge_color):
            pc.set_facecolor(c)
            pc.set_edgecolor(ec)
        errors = False

    elif plot_style == "violin":
        parts = ax.violinplot([s.get_sample() for s in samples_list],
                      vert=True,
                      showmeans=False,
                      showextrema=False,
                      positions=x,
                      )
        ax.set_xticks(x)
        ax.set_xticklabels(labels)
        errors = True
        i=0
        for pc, c, ec in zip(parts['bodies'], color, edge_color):
            pc.set_facecolor(c)
            pc.set_edgecolor(ec)
            if hatch is not None:
                pc.set_hatch(hatch[i])
                i+=1
    elif plot_style == "histogram":
        y_lim = 0
        # for scatter if needed
        y_centers = []
        x_jitter_width = []
        for i, dataset in enumerate([s.get_sample() for s in samples_list]):
            hist, bin_edges = np.histogram(dataset, bins=(np.arange(np.max(dataset) + 2) - 0.5))
            norm_hist = hist/(1.5*np.max(hist))
            bottom = bin_edges[:-1] + 0.5
            heights = np.ones((bin_edges.size - 1))
            lefts = i - 0.5*norm_hist
            y_lim = max(y_lim, np.max(dataset))
            if hatch is not None and hatch[i] is not None:
                ax.barh(bottom, norm_hist, height=heights, left=lefts, color=color[i], edgecolor=edge_color[i],
                        linewidth=1, alpha=0.6, hatch=hatch[i])
            else:
                ax.barh(bottom, norm_hist, height=heights, left=lefts, color=color[i], edgecolor=edge_color[i], linewidth=1, alpha=0.6)
            if show_statistics:
                for j, val in enumerate(norm_hist):
                    ax.text(i, bin_edges[j] + 0.5,
                            "%f.1"%(100*(val/np.sum(norm_hist))),
                            horizontalalignment='center'
                            )
            y_centers.append(bottom)
            x_jitter_width.append(norm_hist)
        errors = True

        ax.set_xlim([-0.4, len(samples_list) - 0.6])
        ax.set_ylim([-1, y_lim+0.6])

    if show_N:
        for index in range(len(samples_list)):
            ax.text(x[index],
                    (averages[index] + standard_errors[index])*1.1,
                    "N = %d" % sample_sizes[index],
                    horizontalalignment='center'
                    )
    if scatter:
        # Adding scattered points on top of the bars
        if plot_style == "violin":

            violin_bodies = parts['bodies']
            paths = [violin_body.get_paths()[0] for violin_body in violin_bodies]
            vertices = [path.vertices for path in paths]

            for i in range(len(x)):
                v = vertices[i]
                left = v[v[:, 0] < x[i]]  # x < center
                right = v[v[:, 0] > x[i]]  # x > center
                increasing_left = np.hstack((1, np.diff(left[:, 1]))) > 1e-7
                f_left = interp1d(left[increasing_left, 1], left[increasing_left, 0], bounds_error=False, fill_value="extrapolate")
                decreasing_right = np.hstack((1, np.diff(right[:, 1]))) < -1e-7
                f_right = interp1d(np.flip(right[decreasing_right, 1]), np.flip(right[decreasing_right, 0]), bounds_error=False, fill_value="extrapolate")
                for group_idx in range(samples_list[i].get_number_of_groups()):
                    group = samples_list[i].get_partial_sample(group_idx)
                    # distribute scatter randomly across the local width of the violin plot
                    x_positions = []
                    if len(group) > 100:
                        group = np.random.choice(group, 100)
                    for y in group:
                        x_min = f_left(y)
                        x_max = f_right(y)
                        jittered_x = np.random.uniform(x_min, x_max)
                        x_positions.append(jittered_x)
                    # ax.scatter(x_positions, group, color=matplotlib.color_sequences["tab10"][group_idx], marker=".", s=10)
                    ax.scatter(x_positions, group, color="gray", marker=".", s=10)
        elif plot_style == "histogram":
            for i in range(len(x)):
                for group_idx in range(samples_list[i].get_number_of_groups()):
                    group = samples_list[i].get_partial_sample(group_idx)
                    # distribute scatter randomly across the local width of the violin plot
                    x_positions = []
                    y_positions = []
                    if len(group) > 100:
                        group = np.random.choice(group, 100)
                    for y in group:
                        w = x_jitter_width[i][np.where(y_centers[i]==y)[0]]
                        jittered_x = np.random.uniform(x[i] - w/2, x[i] + w/2)
                        jittered_y = np.random.uniform(y + 0.3, y - 0.3)
                        x_positions.append(jittered_x)
                        y_positions.append(jittered_y)
                    # ax.scatter(x_positions, y_positions, color=matplotlib.color_sequences["tab10"][group_idx], marker=".", s=10)
                    ax.scatter(x_positions, y_positions, color="gray", marker=".", s=10)
        else:
            y_jitter = 0
            for i in range(len(x)):
                for group_idx in range(samples_list[i].get_number_of_groups()):
                    group = samples_list[i].get_partial_sample(group_idx)
                    # distribute scatter randomly across whole width of bar
                    # ax.scatter(x[i] + np.random.random(group.size) * w - w / 2, group + y_jitter, color=matplotlib.color_sequences["tab10"][group_idx], marker=".", s=10)
                    ax.scatter(x[i] + np.random.random(group.size) * w - w / 2, group + y_jitter,
                               color="gray", marker=".", s=10)

    if errors:
        for i in range(len(x)):
            ax.errorbar(x[i],
                    averages[i],
                    yerr=standard_errors[i],
                    color="black",
                    marker="*",
                    markersize=10,
                    capsize=10,
                    elinewidth=2
                    )
    # Setting plot y-limits
    data_max = np.max(np.array([sample.get_max() for sample in samples_list]))
    data_min = np.min(np.array([sample.get_min() for sample in samples_list]))
    max_error = np.max(np.array([averages[i] + standard_errors[i] for i in range(len(samples_list))]))
    min_error = np.min(np.array([averages[i] - standard_errors[i] for i in range(len(samples_list))]))
    low_lim = min(min_error, data_min)
    high_lim = max(max_error, data_max)
    addition = (high_lim - low_lim) * 0.1
    # ax.set_ylim([low_lim - addition, high_lim + addition])

    # Adding p-value brackets for each compared pair
    if show_statistics:
        heights = np.zeros((len(samples_list),))
        for index, pair in enumerate(pairs_to_compare):
            sample1_index, sample2_index = pair
            dh = max(np.max(heights[sample1_index:sample2_index])/2 +0.1, 0.3)
            heights[sample1_index:sample2_index] = barplot_annotate_brackets(sample1_index, sample2_index, pvalues[index], x,
                                      averages, yerr=standard_errors, fs=10, maxasterix=4, dh=0.1, ax=ax)

    res = {"averages": averages, "standard errors":standard_errors, "sample sizes": sample_sizes,
           "pvalues": {pairs_to_compare[i]: pvalues[i] for i in range(len(pairs_to_compare))}}
    return fig, ax, res


if __name__ == "__main__":
    # Testing methods
    rng = np.random.default_rng()
    samples_list = [sp.stats.norm.rvs(loc=i, scale=(i+1), size=100, random_state=rng) for i in range(5)]
    pairs = [(0,3), (1,4)]
    labels = ['sample%d' % (i+1) for i in range(5)]
    res = compare_and_plot_samples(samples_list, labels, pairs)
    plt.show()
    print(res)



